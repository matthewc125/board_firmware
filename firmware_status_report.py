"""Build colored Tool × board-type firmware status Excel workbooks."""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import db

FILLS = {
    "newest": PatternFill("solid", fgColor="92D050"),  # green
    "middle": PatternFill("solid", fgColor="FFFF00"),  # yellow
    "oldest": PatternFill("solid", fgColor="FF6B6B"),  # red
}

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def build_firmware_status_workbook(matrix: dict | None = None) -> bytes:
    matrix = matrix or db.firmware_status_matrix()
    columns = matrix["columns"]
    rows = matrix["rows"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Firmware Status"

    headers = [""] + [col["label"] for col in columns]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    for row_idx, row in enumerate(rows, start=2):
        tool_cell = ws.cell(row_idx, 1, row["tool"])
        tool_cell.font = HEADER_FONT
        tool_cell.alignment = Alignment(horizontal="left", vertical="center")
        tool_cell.border = THIN

        for col_idx, cell_data in enumerate(row["cells"], start=2):
            value = cell_data.get("firmware") or ""
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN
            rank = cell_data.get("rank")
            if value and rank in FILLS:
                cell.fill = FILLS[rank]

    ws.column_dimensions["A"].width = 10
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def default_filename() -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    return f"firmware_status_{stamp}.xlsx"
