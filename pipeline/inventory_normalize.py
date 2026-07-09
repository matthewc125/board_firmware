"""Normalization helpers for inventory spreadsheet imports."""
from __future__ import annotations

import re
from datetime import date, datetime

import pandas as pd

MANUFACTURER = "PDF Solutions Inc."
NA_VALUES = {"", "N/A", "NAN", "NOT_APPLICABLE", "UNKNOWN", "NONE"}

ETR_ROLE_TO_BOARD_NAME = {
    "LowerSub": "LSF",
    "UpperSub": "USF",
    "Blanker": "Blanker",
    "LowerMain": "ES",
    "UpperMain": "ES",
    "FastFocus": "FF",
    "Stig": "ES",
    "Objective": "LSCC",
}

ETR_ROLE_TO_SLOT = {
    "LowerSub": "4",
    "UpperSub": "5",
    "Blanker": "6",
    "ADC": "ADC",
    "Preamp": "Bench",
    "Backplane": "Bench",
}


def clean_text(value) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if text.upper() in NA_VALUES:
        return None
    return text


INVALID_TOOLS = frozenset({
    "?", "ON DESK", "ON BENCH", "STEVE", "ON DESK", "UNKNOWN", "N/A",
})


def normalize_tool(value) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    lines = [line.strip() for line in re.split(r"[\n\r]+", text) if line.strip()]
    if not lines:
        return None
    current = lines[-1]
    lower = current.lower()
    if lower in ("on bench", "bench"):
        return "Bench"
    if lower in ("steve", "on desk") or "desk" in lower:
        return None
    if current.strip() in ("?", "??") or current.upper() in INVALID_TOOLS:
        return None
    m = re.search(r"tool[- ]?(\d+)", current, re.IGNORECASE)
    if m:
        return f"Tool{m.group(1)}"
    if current.isdigit():
        return f"Tool{current}"
    if current.upper().startswith("TOOL"):
        return re.sub(r"tool[- ]?", "Tool", current, flags=re.IGNORECASE)
    if current.upper() in ("OA", "CTS1", "BENCH"):
        return "Bench" if current.upper() == "BENCH" else current.upper()
    return current


def split_tool_history(value) -> tuple[list[str], str | None]:
    text = clean_text(value)
    if text is None:
        return [], None
    lines = [line.strip() for line in re.split(r"[\n\r]+", text) if line.strip()]
    if not lines:
        return [], None
    normalized = [normalize_tool(line) or line for line in lines]
    return normalized[:-1], normalized[-1]


def parse_part_revision(pn_value) -> tuple[str | None, str | None]:
    text = clean_text(pn_value)
    if text is None:
        return None, None
    m = re.match(r"^(\d{4}-\d{5}-\d{2})(?:\s+Rev\s+(.+))?$", text, re.IGNORECASE)
    if m:
        part = m.group(1)
        rev = m.group(2)
        return part, f"Rev {rev}" if rev else None
    m = re.match(r"^(\d{4}-\d{5}-\d{2})$", text)
    if m:
        return m.group(1), None
    return text, None


def display_serial(value) -> str:
    """Full serial after SN prefix, e.g. SNB0014 -> B0014, SN013 -> 013."""
    text = clean_text(value)
    if not text:
        return ""
    if ":" in text:
        text = text.split(":", 1)[1]
    text = text.upper().replace(" ", "")
    if "?" in text:
        return text.replace("SN", "", 1) if text.startswith("SN") else text
    m = re.match(r"^SN([A-Z]+)(\d+)$", text)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    m = re.match(r"^SN(\d+)$", text)
    if m:
        return m.group(1)
    return text


def short_serial(value) -> str:
    """Display serial for boards table (preserves type letter after SN)."""
    displayed = display_serial(value)
    return displayed or (clean_text(value) or "")


def serial_suffix(value) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    displayed = display_serial(text)
    m = re.search(r"(\d+)$", displayed)
    if m:
        return str(int(m.group(1)))
    return displayed or text


def normalize_inventory_serial(value) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    return text.upper().replace(" ", "")


def strong_match_key(part_number: str | None, inventory_serial: str | None) -> str | None:
    part = clean_text(part_number)
    inv = normalize_inventory_serial(inventory_serial)
    if not part or not inv:
        return None
    return f"{part}|{inv}"


def pico_match_key(part_number: str | None, inventory_serial: str | None) -> str | None:
    part = clean_text(part_number)
    suffix = serial_suffix(inventory_serial)
    if not part or not suffix:
        return None
    return f"{part}|{suffix}"


def format_revision(value) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    if text.lower().startswith("rev "):
        return text
    return f"Rev {text}"


def split_datetime(value) -> tuple[str | None, str | None]:
    if isinstance(value, datetime):
        return value.date().isoformat(), value.strftime("%H:%M:%S")
    if isinstance(value, date):
        return value.isoformat(), None
    text = clean_text(value)
    if text is None:
        return None, None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date().isoformat(), parsed.strftime("%H:%M:%S") if " " in fmt else None
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(value)
        if pd.isna(parsed):
            return None, None
        return parsed.date().isoformat(), parsed.strftime("%H:%M:%S")
    except (ValueError, TypeError):
        return None, None


def map_etr_type(etr_type: str, role: str | None) -> tuple[str, str, str]:
    etr_type = (etr_type or "").strip()
    role = clean_text(role)
    if etr_type == "EM":
        return "EM1", "LSCC", "Bench"
    if etr_type == "ES":
        board_name = ETR_ROLE_TO_BOARD_NAME.get(role or "", "ES")
        slot = ETR_ROLE_TO_SLOT.get(role or "", "Bench")
        return "ES4", board_name, slot
    if etr_type == "Obj":
        return "OBJ", "LSCC", "Bench"
    if etr_type == "Preamp":
        return "Pico", "Preamp", "Bench"
    if etr_type == "ADC":
        return "ADC", "ADC", "Bench"
    if etr_type in ("Backplane", "Flex", "Interposer", "FF", "Extender"):
        return "Column", etr_type, "Bench"
    return etr_type or "Unknown", etr_type or "Unknown", "Bench"


def map_pico_board(pico_board: str) -> tuple[str, str]:
    pico_board = (pico_board or "").strip()
    if pico_board in ("Dual Pico", "Dual Pico2", "Single Pico"):
        return "Pico", pico_board
    if pico_board in ("Preamp", "Preamp2"):
        return "Pico", pico_board
    if pico_board == "Blanker":
        return "Pico", "Blanker"
    if pico_board in ("BAP", "BAP2"):
        return "BAP", pico_board
    return pico_board, pico_board


def classify_event_type(description: str) -> str:
    lower = description.lower()
    if any(k in lower for k in ("install", "swapped onto", "on tool")):
        return "install"
    if any(k in lower for k in ("removed", "took to", "ship")):
        return "remove"
    if any(k in lower for k in ("test", "retest", "working")):
        return "test"
    if any(k in lower for k in ("received", "sent to", "brought")):
        return "receive"
    if any(k in lower for k in ("repair", "rework", "clean")):
        return "repair"
    if "bench" in lower or "location" in lower:
        return "location"
    return "note"


def is_scrapped_status(status: str | None) -> bool:
    text = clean_text(status)
    return text is not None and text.lower() == "scrapped"


def load_strikethrough_keys(etr_path: str | None, pico_path: str | None) -> set[str]:
    """Return normalized inventory serial keys for strikethrough rows in source spreadsheets."""
    from openpyxl import load_workbook

    keys: set[str] = set()
    if etr_path:
        wb = load_workbook(etr_path, data_only=True)
        ws = wb.active
        headers = [clean_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row):
                continue
            strike = any(c.font and c.font.strike for c in row if c.value is not None)
            if not strike:
                continue
            etr_type = clean_text(row[col["Type"]].value if "Type" in col else None) or "Unknown"
            sn = clean_text(row[col["SN"]].value if "SN" in col else None)
            if sn:
                keys.add(normalize_inventory_serial(f"{etr_type}:{sn}") or "")
                keys.add(normalize_inventory_serial(sn) or "")
        wb.close()
    if pico_path:
        wb = load_workbook(pico_path, data_only=True)
        ws = wb.active
        headers = [clean_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row):
                continue
            strike = any(c.font and c.font.strike for c in row if c.value is not None)
            if not strike:
                continue
            sn = clean_text(row[col["SN"]].value if "SN" in col else None)
            if sn:
                keys.add(normalize_inventory_serial(sn) or "")
        wb.close()
    keys.discard("")
    return keys


def board_matches_strikethrough(inventory_serial: str | None, strike_keys: set[str]) -> bool:
    inv = normalize_inventory_serial(inventory_serial)
    if not inv or not strike_keys:
        return False
    if inv in strike_keys:
        return True
    if ":" in inv:
        suffix = inv.split(":", 1)[1]
        return suffix in strike_keys
    return False


def etr_row_excluded(row, strike_keys: set[str]) -> str | None:
    etr_type = clean_text(row.get("Type")) or "Unknown"
    sn = clean_text(row.get("SN"))
    inv = normalize_inventory_serial(f"{etr_type}:{sn}") if sn else None
    if inv and board_matches_strikethrough(inv, strike_keys):
        return "strikethrough"
    if is_scrapped_status(row.get("Status")):
        return "scrapped"
    if not clean_text(row.get("Firmware")):
        return "no_firmware"
    return None


def pico_row_excluded(row, strike_keys: set[str]) -> str | None:
    sn = clean_text(row.get("SN"))
    if sn and board_matches_strikethrough(sn, strike_keys):
        return "strikethrough"
    return None


def board_removal_reasons(
    conn,
    board: dict,
    strike_keys: set[str],
) -> list[str]:
    reasons: list[str] = []
    if board_matches_strikethrough(board.get("inventory_serial"), strike_keys):
        reasons.append("strikethrough")
    if is_scrapped_status(board.get("status")):
        reasons.append("scrapped")
    fw_count = conn.execute(
        "SELECT COUNT(*) FROM firmware_history WHERE board_id = ?",
        (board["board_id"],),
    ).fetchone()[0]
    if fw_count == 0:
        reasons.append("no_firmware")
    return reasons


def parse_pico_status_events(status_text, board_label: str, sn: str) -> list[dict]:
    text = clean_text(status_text)
    if not text:
        return []
    events = []
    for match in re.finditer(r"\[(\d{2}/\d{2}/\d{4})\]\s*([^\[]+)", text):
        event_date, _ = split_datetime(match.group(1))
        if not event_date:
            continue
        description = match.group(2).strip()
        tool = None
        tool_match = re.search(r"tool[- ]?(\d+)", description, re.IGNORECASE)
        if tool_match:
            tool = f"Tool{tool_match.group(1)}"
        events.append({
            "event_date": event_date,
            "event_time": None,
            "event_type": classify_event_type(description),
            "description": description,
            "tool": tool,
            "source": "pico_status",
            "source_ref": f"{board_label}:{sn}:{event_date}",
        })
    return events
